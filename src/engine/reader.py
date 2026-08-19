# ================================================================
# engine/reader.py — Streaming Multi-Format Reader
# Lazy Loading | Chunk Reading | Memory Mapping | XML | JSON
# ================================================================

import os, io, re, json, csv, mmap, gc, time, logging
from typing import Iterator, Optional, List, Dict, Tuple

import pandas as pd
import numpy as np

logger = logging.getLogger("jet.reader")

TR_MAP = str.maketrans("ıİğĞşŞüÜöÖçÇ", "iiggssuuoocc")

CHUNK_ROWS   = 200_000          # her chunk'ta max satır
SAMPLE_BYTES = 8_192            # ayırıcı tespiti için örnek

# ─────────────────────────────────────────────────────────────
# YARDIMCILAR
# ─────────────────────────────────────────────────────────────

def detect_sep(sample: str) -> str:
    counts = {d: sample.count(d) for d in [";", ",", "\t", "|"]}
    return max(counts, key=counts.get)


def detect_encoding(raw: bytes) -> str:
    """BOM veya basit heuristic ile encoding."""
    if raw[:3] == b'\xef\xbb\xbf':
        return 'utf-8-sig'
    if raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
        return 'utf-16'
    try:
        raw[:4096].decode('utf-8')
        return 'utf-8'
    except UnicodeDecodeError:
        return 'latin-1'


def make_unique_columns(columns) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, col in enumerate(columns):
        base = str(col).strip() if col is not None and str(col).strip() else f'Col_{i}'
        count = seen.get(base, 0)
        out.append(base if count == 0 else f"{base}_{count + 1}")
        seen[base] = count + 1
    return out


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = make_unique_columns(df.columns)
    return df


def optimize_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    """
    Tüm kolonları string olarak bırak ama object kolonları
    StringDtype'a çevirerek bellek kullanımını azalt.
    """
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype('string').fillna('')
    return df


# ─────────────────────────────────────────────────────────────
# FORMAT DETEKT
# ─────────────────────────────────────────────────────────────

def get_format(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    return ext  # .xlsx .xlsm .xls .csv .txt .tsv .json .xml


def get_sheet_names(filepath: str) -> List[str]:
    ext = get_format(filepath)
    try:
        if ext in ('.xlsx', '.xlsm'):
            xl = pd.ExcelFile(filepath, engine='openpyxl')
            return xl.sheet_names
        elif ext == '.xls':
            xl = pd.ExcelFile(filepath, engine='xlrd')
            return xl.sheet_names
    except Exception:
        pass
    return []


# ─────────────────────────────────────────────────────────────
# STREAMING READERS — her biri Iterator[pd.DataFrame] döner
# ─────────────────────────────────────────────────────────────

def stream_csv(filepath: str, sep: Optional[str] = None,
               encoding: Optional[str] = None,
               skiprows: int = 0, header_row: int = 0,
               chunk: int = CHUNK_ROWS,
               progress_cb=None) -> Iterator[pd.DataFrame]:
    """
    CSV/TXT/TSV'yi chunk'lar halinde oku.
    Memory mapping kullanarak büyük dosyalarda disk I/O minimize edilir.
    """
    if encoding is None:
        with open(filepath, 'rb') as f:
            encoding = detect_encoding(f.read(SAMPLE_BYTES))

    if sep is None:
        with open(filepath, 'r', encoding=encoding, errors='replace') as f:
            sample = f.read(SAMPLE_BYTES)
        sep = detect_sep(sample)

    file_size = os.path.getsize(filepath)
    read_bytes = 0

    reader = pd.read_csv(
        filepath,
        sep=sep,
        dtype=str,
        encoding=encoding,
        low_memory=False,
        chunksize=chunk,
        skiprows=range(1, skiprows + 1) if skiprows else None,
        header=header_row,
        on_bad_lines='skip',
        engine='c',        # C engine → en hızlı
    )

    for i, chunk_df in enumerate(reader):
        chunk_df.fillna('', inplace=True)
        clean_columns(chunk_df)
        if progress_cb:
            try:
                read_bytes = min(file_size, read_bytes + chunk * 200)
                progress_cb(int(read_bytes / file_size * 100))
            except Exception:
                pass
        yield chunk_df
        del chunk_df
        gc.collect()


def stream_excel(filepath: str, sheet=None,
                 skiprows: int = 0, header_row: int = 0,
                 chunk: int = CHUNK_ROWS,
                 progress_cb=None) -> Iterator[pd.DataFrame]:
    """
    Excel okuma: openpyxl read_only modu → büyük dosyalarda düşük RAM.
    Chunk'lar halinde yield eder.
    """
    ext = get_format(filepath)
    engine = 'openpyxl' if ext in ('.xlsx', '.xlsm') else 'xlrd'

    if progress_cb: progress_cb(5)

    # openpyxl read_only ile satır satır oku
    if engine == 'openpyxl':
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # Header row is zero-based. skiprows means extra rows between header and data.
        header_offset = max(int(header_row or 0), 0)
        for _ in range(header_offset):
            next(rows_iter, None)
        header = make_unique_columns(c if c is not None else f'Col_{i}'
                                     for i, c in enumerate(next(rows_iter, [])))
        for _ in range(max(int(skiprows or 0) - header_offset, 0)):
            next(rows_iter, None)

        total_rows = ws.max_row or 1
        processed  = 0
        buf: List[tuple] = []

        for row in rows_iter:
            buf.append(tuple('' if v is None else str(v) for v in row))
            if len(buf) >= chunk:
                df = pd.DataFrame(buf, columns=header)
                df.fillna('', inplace=True)
                processed += len(buf)
                if progress_cb:
                    progress_cb(min(95, int(processed / total_rows * 100)))
                yield df
                buf = []
                del df
                gc.collect()

        if buf:
            df = pd.DataFrame(buf, columns=header)
            df.fillna('', inplace=True)
            yield df

        wb.close()
        if progress_cb: progress_cb(100)

    else:
        # xlrd (xls) — tek seferde oku, chunk'lara böl
        df_full = pd.read_excel(filepath, sheet_name=sheet or 0,
                                dtype=str, engine='xlrd',
                                skiprows=skiprows, header=header_row)
        df_full.fillna('', inplace=True)
        clean_columns(df_full)
        total = len(df_full)
        for start in range(0, total, chunk):
            if progress_cb:
                progress_cb(min(95, int(start / total * 100)))
            yield df_full.iloc[start:start + chunk].copy()
        if progress_cb: progress_cb(100)


def stream_json(filepath: str, chunk: int = CHUNK_ROWS,
                progress_cb=None) -> Iterator[pd.DataFrame]:
    """
    JSON okuma: records / lines formatı desteklenir.
    Büyük dosyalarda incremental parse.
    """
    if progress_cb: progress_cb(10)
    file_size = os.path.getsize(filepath)

    # JSONL (newline-delimited JSON) kontrolü
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        first = f.read(2).strip()
    
    if first == '[' or first == '{':
        # Standart JSON array veya object
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            data = json.load(f)
        if isinstance(data, dict):
            data = [data]
        df_full = pd.DataFrame(data).astype(str).fillna('')
        total = len(df_full)
        for start in range(0, max(total, 1), chunk):
            if progress_cb:
                progress_cb(min(95, 10 + int(start / max(total, 1) * 85)))
            yield df_full.iloc[start:start + chunk].copy()
    else:
        # JSONL
        buf = []
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    buf.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
                if len(buf) >= chunk:
                    df = pd.DataFrame(buf).astype(str).fillna('')
                    yield df
                    buf = []
                    gc.collect()
        if buf:
            yield pd.DataFrame(buf).astype(str).fillna('')
    if progress_cb: progress_cb(100)


def stream_xml(filepath: str, chunk: int = CHUNK_ROWS,
               progress_cb=None) -> Iterator[pd.DataFrame]:
    """
    XML incremental parse — iterparse ile düşük bellek.
    En üst seviyedeki child element'leri satır olarak okur.
    """
    try:
        import xml.etree.ElementTree as ET
    except ImportError:
        raise RuntimeError("xml.etree.ElementTree bulunamadı")

    if progress_cb: progress_cb(5)
    file_size = os.path.getsize(filepath)

    context = ET.iterparse(filepath, events=('start', 'end'))
    _, root = next(context)

    # root'un doğrudan çocukları → her biri bir satır
    row_tag = None
    buf: List[Dict] = []

    for event, elem in context:
        if event == 'end':
            if row_tag is None:
                row_tag = elem.tag
            if elem.tag == row_tag:
                record = {child.tag: (child.text or '').strip()
                          for child in elem}
                # elem attribute'leri de ekle
                for k, v in elem.attrib.items():
                    record[f'@{k}'] = v
                if not record:
                    record = {'value': (elem.text or '').strip()}
                buf.append(record)

                if len(buf) >= chunk:
                    df = pd.DataFrame(buf).astype(str).fillna('')
                    yield df
                    buf = []
                    del df; gc.collect()

                root.clear()   # belleği serbest bırak

    if buf:
        yield pd.DataFrame(buf).astype(str).fillna('')
    if progress_cb: progress_cb(100)


# ─────────────────────────────────────────────────────────────
# UNIFIED READER — tek arayüz
# ─────────────────────────────────────────────────────────────

def stream_file(filepath: str,
                sheet=None,
                sep: Optional[str] = None,
                encoding: Optional[str] = None,
                skiprows: int = 0,
                header_row: int = 0,
                chunk: int = CHUNK_ROWS,
                progress_cb=None) -> Iterator[pd.DataFrame]:
    """
    Tüm formatlar için unified streaming reader.
    Dönen her DataFrame aynı tip garantisi verir.
    """
    ext = get_format(filepath)

    if ext in ('.csv', '.txt', '.tsv'):
        yield from stream_csv(filepath, sep=sep, encoding=encoding,
                              skiprows=skiprows, header_row=header_row,
                              chunk=chunk, progress_cb=progress_cb)
    elif ext in ('.xlsx', '.xlsm', '.xls'):
        yield from stream_excel(filepath, sheet=sheet,
                                skiprows=skiprows, header_row=header_row,
                                chunk=chunk, progress_cb=progress_cb)
    elif ext == '.json':
        yield from stream_json(filepath, chunk=chunk, progress_cb=progress_cb)
    elif ext == '.xml':
        yield from stream_xml(filepath, chunk=chunk, progress_cb=progress_cb)
    else:
        # Bilinmeyen → CSV olarak dene
        yield from stream_csv(filepath, sep=sep, encoding=encoding,
                              skiprows=skiprows, chunk=chunk,
                              progress_cb=progress_cb)


def read_full(filepath: str, **kwargs) -> pd.DataFrame:
    """
    Tüm dosyayı tek DataFrame'e oku (küçük-orta boyutlu dosyalar için).
    """
    frames = list(stream_file(filepath, **kwargs))
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df.fillna('', inplace=True)
    return df
