# ================================================================
# engine/analyses.py — JET Analiz Fonksiyonları
# Kullanıcı | Kelime | Unusual — chunk bazlı, paralel, progress-aware
# ================================================================

import os, re, gc, shutil, logging, time
from typing import List, Dict, Optional, Callable, Any, Tuple

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .executor import ProgressReporter

logger = logging.getLogger("jet.analyses")

TR_MAP = str.maketrans("ıİğĞşŞüÜöÖçÇ", "iiggssuuoocc")

# ─────────────────────────────────────────────────────────────
# YARDIMCILAR
# ─────────────────────────────────────────────────────────────

def _norm(text: str) -> str:
    if not isinstance(text, str):
        text = str(text)
    return text.translate(TR_MAP).lower()


def _word_pat(kw: str) -> re.Pattern:
    nk = re.escape(_norm(kw))
    return re.compile(r'(?<![a-z0-9])' + nk + r'(?![a-z0-9])')


def _safe(v) -> Any:
    if isinstance(v, (np.integer,)):  return int(v)
    if isinstance(v, (np.floating,)): return float(v)
    if isinstance(v, (np.bool_,)):    return bool(v)
    try:
        if pd.isna(v): return ''
    except Exception:
        pass
    return str(v) if not isinstance(v, str) else v


def _safe_output_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        s = out[col]
        if isinstance(s, pd.DataFrame):
            continue
        try:
            if pd.api.types.is_numeric_dtype(s):
                out[col] = s.replace([np.inf, -np.inf], np.nan).fillna(0)
            elif pd.api.types.is_bool_dtype(s):
                out[col] = s.fillna(False)
            elif pd.api.types.is_datetime64_any_dtype(s):
                out[col] = s
            else:
                out[col] = s.astype("object").where(pd.notna(s), "")
        except Exception:
            logger.warning(f"Kolon temizleme atlandi: {col}", exc_info=True)
    return out


def _fmt_ws(ws, n_cols: int, hcol: str = "4A4393"):
    hf    = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    df_   = Font(name="Arial", size=9)
    ca    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hfill = PatternFill(start_color=hcol, end_color=hcol, fill_type="solid")
    for cell in ws[1]:
        cell.font = hf; cell.alignment = ca; cell.fill = hfill
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.font = df_
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.column_dimensions["A"].width = 30
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24


def _write_excel(df: pd.DataFrame, path: str,
                 sheet: str = "Sayfa1", hcol: str = "4A4393"):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name=sheet, index=False)
        _fmt_ws(w.sheets[sheet], len(df.columns), hcol)


def _get_output_folder(name: str = "JET_Ciktilari") -> str:
    p = os.path.join(os.path.expanduser("~"), "Desktop", name)
    os.makedirs(p, exist_ok=True)
    return p


def _paste_to_template(tpl_path: str, sheet_name: str,
                       df: pd.DataFrame, start_row: int = 52,
                       start_col: int = 2) -> Tuple[bool, str]:
    try:
        wb = load_workbook(tpl_path, keep_vba=True)
    except Exception as e:
        return False, str(e)
    if sheet_name not in wb.sheetnames:
        return False, f"'{sheet_name}' sheet yok. Mevcut: {wb.sheetnames}"
    ws  = wb[sheet_name]
    hf  = Font(name="Arial", bold=True, size=10)
    df_ = Font(name="Arial", size=9)
    ca  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, cn in enumerate(df.columns, start_col):
        c = ws.cell(row=start_row - 1, column=ci, value=cn)
        c.font = hf; c.alignment = ca
    for ri, row in enumerate(df.itertuples(index=False), start_row):
        for ci, val in enumerate(row, start_col):
            ws.cell(row=ri, column=ci, value=_safe(val)).font = df_
    wb.save(tpl_path)
    return True, "OK"


def _template_helper(tpl_path: str, out_folder: str,
                     sheet_name: str, df: pd.DataFrame) -> str:
    if not tpl_path or not os.path.exists(tpl_path):
        return ''
    tpl_dest = os.path.join(out_folder, 'JET_Template_Doldurulmus.xlsm')
    if not os.path.exists(tpl_dest):
        shutil.copy2(tpl_path, tpl_dest)
    ok, msg = _paste_to_template(tpl_dest, sheet_name, df)
    return 'Template yapıştırıldı.' if ok else f'Template hatası: {msg}'


def _df_preview(df: pd.DataFrame, n: int = 200) -> Tuple[List[str], List[dict]]:
    cols = list(df.columns)
    rows = [{c: _safe(v) for c, v in zip(cols, row)}
            for row in df.head(n).itertuples(index=False)]
    return cols, rows


# ─────────────────────────────────────────────────────────────
# KULLANICI TARAMASI
# ─────────────────────────────────────────────────────────────

def run_kullanici(df: pd.DataFrame,
                  user_col: str,
                  borc_col: Optional[str],
                  alacak_col: Optional[str],
                  tutar_kols: List[str],
                  output_folder: str,
                  template_path: str,
                  _progress_cb: Optional[Callable] = None) -> dict:
    """
    Kullanıcı taraması — groupby + aggregation.
    Büyük df'lerde chunked aggregation ile düşük bellek.
    """
    pr = ProgressReporter(_progress_cb, total_rows=len(df))
    pr.set(5, "Kolonlar hazırlanıyor...")

    agg_cols = [user_col]
    if borc_col:    agg_cols.append(borc_col)
    if alacak_col:  agg_cols.append(alacak_col)
    agg_cols += [c for c in tutar_kols if c not in agg_cols]

    df_num = df[agg_cols].copy()
    for c in agg_cols[1:]:
        df_num[c] = pd.to_numeric(
            df_num[c].astype(str).str.replace(',', '.', regex=False),
            errors='coerce'
        ).fillna(0)

    pr.set(30, "Gruplama yapılıyor...")

    agg_dict: Dict[str, tuple] = {'NO_OF_RECORDS': (user_col, 'count')}
    if borc_col:   agg_dict['BORC_TUTARI_SUM']   = (borc_col,   'sum')
    if alacak_col: agg_dict['ALACAK_TUTARI_SUM'] = (alacak_col, 'sum')
    for c in tutar_kols:
        if c not in [borc_col, alacak_col]:
            agg_dict[f'{c}_SUM'] = (c, 'sum')

    result_df = (
        df_num.groupby(user_col, sort=False)
              .agg(**agg_dict).reset_index()
              .sort_values('NO_OF_RECORDS', ascending=False)
    )
    result_df.rename(columns={user_col: 'KULLANICI_KIMLIGI'}, inplace=True)

    pr.set(70, "Excel yazılıyor...")

    out      = _get_output_folder(output_folder)
    out_path = os.path.join(out, 'Kullanici_Taramasi.xlsx')
    _write_excel(result_df, out_path, 'Kullanıcı Taraması', '534AB7')

    tpl_msg  = _template_helper(template_path, out,
                                'Step 9 JE Testing - Kullanıcı', result_df)
    cols, rows = _df_preview(result_df, 200)

    pr.done("Kullanıcı taraması tamamlandı")
    return {
        'ok': True, 'count': len(result_df),
        'savedTo': out_path, 'tplMsg': tpl_msg,
        'columns': cols, 'preview': rows,
    }


# ─────────────────────────────────────────────────────────────
# KELİME TARAMASI — Chunk + Paralel
# ─────────────────────────────────────────────────────────────

def _scan_chunk(chunk_df: pd.DataFrame,
                scan_cols: List[str],
                patterns: Dict[str, re.Pattern],
                norm_map: Optional[Dict[str, pd.Series]] = None
                ) -> Tuple[List[dict], Dict[str, int]]:
    """Tek chunk üzerinde kelime taraması (paralel çağrılabilir)."""
    kw_counts  = {kw: 0 for kw in patterns}
    results    = []
    orig_cols  = list(chunk_df.columns)

    # Normalize
    nc = {c: chunk_df[c].map(_norm) for c in scan_cols if c in chunk_df.columns}

    # Vektörel ön filtre
    global_mask = pd.Series(False, index=chunk_df.index)
    for pat in patterns.values():
        for c in nc:
            global_mask |= nc[c].str.contains(pat, na=False, regex=True)

    for idx in chunk_df.index[global_mask]:
        found = []
        for kw, pat in patterns.items():
            for c in nc:
                if pat.search(nc[c].at[idx]):
                    found.append(kw); break
        if found:
            row = {col: chunk_df.at[idx, col] for col in orig_cols if col in chunk_df.columns}
            row['Eşleşen Anahtar Kelimeler'] = ', '.join(found)
            results.append(row)
            for kw in found:
                kw_counts[kw] += 1

    return results, kw_counts


def run_kelime(df: pd.DataFrame,
               scan_cols: List[str],
               keywords: List[str],
               scan_mode: str,
               output_folder: str,
               template_path: str,
               chunk_size: int = 500_000,
               _progress_cb: Optional[Callable] = None) -> dict:
    """
    Kelime taraması — büyük df'lerde chunk bazlı tarama.
    """
    pr = ProgressReporter(_progress_cb, total_rows=len(df))
    pr.set(5, "Patternler hazırlanıyor...")

    # Pattern derle
    patterns: Dict[str, re.Pattern] = {}
    for kw in keywords:
        nk = _norm(kw)
        patterns[kw] = (re.compile(re.escape(nk), re.IGNORECASE)
                        if scan_mode == 'icerik' else _word_pat(kw))

    all_results:  List[dict] = []
    total_counts: Dict[str, int] = {kw: 0 for kw in keywords}

    total = len(df)
    orig_cols = list(df.columns)

    for start in range(0, total, chunk_size):
        chunk = df.iloc[start:start + chunk_size]
        results, counts = _scan_chunk(chunk, scan_cols, patterns)
        all_results.extend(results)
        for kw, n in counts.items():
            total_counts[kw] += n

        done = min(start + chunk_size, total)
        pct  = int(done / total * 85) + 5
        pr.set(pct, f"Taranıyor... {done:,}/{total:,}")
        del chunk; gc.collect()

    pr.set(92, "Excel yazılıyor...")

    out      = _get_output_folder(output_folder)
    saved_to = ''
    kdf      = pd.DataFrame()

    if all_results:
        kdf = pd.DataFrame(all_results)[
            ['Eşleşen Anahtar Kelimeler'] + orig_cols
        ]
        saved_to = os.path.join(out, 'Kelime_Taramasi.xlsx')

        with pd.ExcelWriter(saved_to, engine='openpyxl') as w:
            kdf.to_excel(w, sheet_name='Kelime Taraması', index=False)
            _fmt_ws(w.sheets['Kelime Taraması'], len(kdf.columns), '1D6B8A')
            # Özet sheet
            ozm  = pd.DataFrame({
                'Bilgi':  ['Mod', 'Toplam Satır', 'Eşleşen Satır', 'Kolonlar'],
                'Değer':  ['TAM KELİME' if scan_mode == 'tam' else 'İÇERİK',
                           total, len(all_results), ', '.join(scan_cols)]
            })
            ozkw = pd.DataFrame({
                'Kelime': list(total_counts.keys()),
                'Sayı':   list(total_counts.values()),
            })
            ozm.to_excel(w,  sheet_name='Özet', index=False, startrow=0)
            ozkw.to_excel(w, sheet_name='Özet', index=False, startrow=len(ozm) + 2)

    tpl_msg = ''
    if all_results:
        tpl_msg = _template_helper(template_path, out,
                                   'Step 9 JE Testing - Kelime', kdf)

    preview = [{k: str(v) for k, v in r.items()} for r in all_results[:200]]
    pr.done("Kelime taraması tamamlandı")

    return {
        'ok':      True,
        'matched': len(all_results),
        'kwCounts':total_counts,
        'savedTo': saved_to,
        'tplMsg':  tpl_msg,
        'preview': preview,
        'columns': ['Eşleşen Anahtar Kelimeler'] + orig_cols,
    }


# ─────────────────────────────────────────────────────────────
# UNUSUAL ANALİZİ — Chunk + Çoklu Kriter
# ─────────────────────────────────────────────────────────────

def run_unusual(df: pd.DataFrame,
                hesap_col: str,
                grup_col: str,
                criteria: List[Dict],
                match_len: int,
                output_folder: str,
                template_path: str,
                _progress_cb: Optional[Callable] = None) -> dict:
    """
    Unusual analizi — filtrelenmiş vektörel hesap kodu eşleştirme.
    """
    pr = ProgressReporter(_progress_cb, total_rows=len(df))
    pr.set(5, "Hesap kodları hazırlanıyor...")

    hesap_s = df[hesap_col].astype(str).str.strip().str[:match_len]

    # Önizleme veya çoklu kriter birleştirme için kullanılacak değişkenler
    toplam_unusual_n = 0
    df_unusual_list = []
    fis_grup_list = []
    kriter_ozet = []
    preview_rows = []

    pr.set(20, "Senaryolar işleniyor...")

    for idx, cr in enumerate(criteria):
        ana_kodlar = [k.strip() for k in cr.get('anaKodlar', []) if k.strip()]
        karsi_kodlar = [k.strip() for k in cr.get('karsiKodlar', []) if k.strip()]

        if not ana_kodlar or not karsi_kodlar:
            continue

        ana_set = set(ana_kodlar)
        karsi_set = set(karsi_kodlar)

        # Sadece ilgili satırları filtrele (performans artışı)
        filtre_mask = hesap_s.isin(ana_set) | hesap_s.isin(karsi_set)
        df_filtre = df.loc[filtre_mask].copy()

        if df_filtre.empty:
            kriter_ozet.append({
                'name': cr['name'],
                'unusual': 0,
                'ana': ana_kodlar,
                'karsi': karsi_kodlar,
            })
            continue

        flag_ana_adi = f"FLAG_{cr['name']}_ANA"
        flag_karsi_adi = f"FLAG_{cr['name']}_KARSI"

        hesap_s_f = df_filtre[hesap_col].astype(str).str.strip().str[:match_len]
        df_filtre[flag_ana_adi] = hesap_s_f.isin(ana_set).astype("int8")
        df_filtre[flag_karsi_adi] = hesap_s_f.isin(karsi_set).astype("int8")

        # Fiş bazında gruplama
        fis_grup = (
            df_filtre.groupby(grup_col, sort=False)[[flag_ana_adi, flag_karsi_adi]]
                     .sum()
                     .reset_index()
        )

        # Unusual kolonu hesaplama
        fn_u = f"UNUSUAL_{cr['name']}"
        fis_grup[fn_u] = (
            (fis_grup[flag_ana_adi] > 0) & (fis_grup[flag_karsi_adi] == 0)
        ).astype("int8")

        unusual_fis_n = int(fis_grup[fn_u].sum())
        toplam_unusual_n += unusual_fis_n

        # Fiş Kayıt sayısı (Orijinal df üzerinden alınıyor)
        fis_kayit = df.groupby(grup_col, sort=False).size().reset_index(name="NO_OF_RECORDS")

        unusual_detay = (
            fis_grup.loc[fis_grup[fn_u] == 1, [grup_col, flag_ana_adi, flag_karsi_adi, fn_u]]
                    .merge(fis_kayit, on=grup_col, how="left")
        )

        df_unusual = df.merge(
            unusual_detay[[grup_col, "NO_OF_RECORDS", flag_ana_adi, flag_karsi_adi, fn_u]],
            on=grup_col,
            how="inner"
        )

        if not df_unusual.empty:
            diger_kols = [c for c in df.columns if c != grup_col]
            kolon_sirasi = ([grup_col, "NO_OF_RECORDS", flag_ana_adi, flag_karsi_adi, fn_u] + diger_kols)
            df_unusual = df_unusual[kolon_sirasi]
            df_unusual_list.append(df_unusual)

            # Preview İçin Hazırlık (Her kriterden max 10 tane al)
            top_fis = unusual_detay.head(10)
            for _, row in top_fis.iterrows():
                preview_rows.append({
                    'fis': str(row[grup_col]),
                    'kayit': int(row['NO_OF_RECORDS']),
                    'unusual': True,
                    'kriterler': [{
                        'name': cr['name'],
                        'unusual': True,
                        'ana': int(row[flag_ana_adi]),
                        'karsi': int(row[flag_karsi_adi]),
                    }]
                })

        fis_grup_list.append(fis_grup)

        kriter_ozet.append({
            'name': cr['name'],
            'unusual': unusual_fis_n,
            'ana': ana_kodlar,
            'karsi': karsi_kodlar,
        })

    pr.set(70, "Sonuçlar birleştiriliyor...")

    # Sonuçları birleştir
    if df_unusual_list:
        final_df_unusual = pd.concat(df_unusual_list, ignore_index=True)
        # Sütunları NaN olanları boş string ile doldur (farklı kriterlerde olmayan bayraklar için)
        final_df_unusual = _safe_output_df(final_df_unusual)
    else:
        final_df_unusual = pd.DataFrame()

    out = _get_output_folder(output_folder)
    out_path = os.path.join(out, 'Unusual_Analiz.xlsx')

    pr.set(85, "Excel yazılıyor...")

    with pd.ExcelWriter(out_path, engine='openpyxl') as w:
        if not final_df_unusual.empty:
            final_df_unusual.to_excel(w, sheet_name='Unusual Detay', index=False)
            _fmt_ws(w.sheets['Unusual Detay'], len(final_df_unusual.columns), '8B2020')
        else:
            pd.DataFrame({'Bilgi': ['Unusual fiş bulunamadı.']}).to_excel(w, sheet_name='Unusual Detay', index=False)

        ozet_df = pd.DataFrame(kriter_ozet)
        ozet_df.to_excel(w, sheet_name='Kriter Özeti', index=False)
        _fmt_ws(w.sheets['Kriter Özeti'], len(ozet_df.columns), '534AB7')

    tpl_msg = ''
    if not final_df_unusual.empty:
        tpl_msg = _template_helper(template_path, out, 'Step 9 JE Testing - Unusual', final_df_unusual)

    pr.done("Unusual analizi tamamlandı")

    return {
        'ok': True,
        'unusualFis': toplam_unusual_n,
        'unusualRows': len(final_df_unusual),
        'totalFis': df[grup_col].nunique(),
        'savedTo': out_path,
        'tplMsg': tpl_msg,
        'preview': preview_rows,
        'kriterOzet': kriter_ozet,
        'criteriaNames': [cr['name'] for cr in criteria],
    }
